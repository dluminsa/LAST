import pandas as pd 
import streamlit as st 
import os
import numpy as np
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time
from pathlib import Path
from streamlit_gsheets import GSheetsConnection
from datetime import datetime

st.set_page_config(
    page_title = 'MPIGI MOCK TX CURR',
    page_icon =":bar_chart"
    )

#st.header('CODE UNDER MAINTENANCE, TRY AGAIN TOMORROW')
#st.stop()
st.subheader('MOCK UP TX CURR AND VL COV')
st.success('WELCOME, this app was developed by Dr. Luminsa Desire, for any concern, reach out to him at desireluminsa@gmail.com')
today = datetime.now()
today = today.strftime("%Y-%m-%d %H:%M")
st.write(f"CURRENT DATE:    {today}")

st.markdown('**FIRST RENAME THESE COLUMNS BEFORE YOU PROCEED:**')
col1, col2, col3 = st.columns([1,1,1])
col1.markdown('Rename the **HIV CLINIC NO.** column to **A**')
col1.markdown('Rename the **ART START DATE** column to **AS**')
col2.markdown('Rename the **RETURN VISIT DATE** column to **RD**')
col2.markdown('Rename the **TRANSFER OUT DATE** column to **TO**')
col3.markdown('Rename the **TRANSFER IN DATE** column to **TI**')
col3.markdown('Rename the **HIV VIRAL LOAD DATE** column to **VD**')
col3.markdown('Rename the **DEATH DATE** column to **DD**')

file = st.file_uploader("Upload your EMR extract here", type=['csv', 'xlsx', 'txt'])
if 'submited' not in st.session_state:
    st.session_state.submited =False
ext = None
if file is not None:
    # Get the file name
    fileN = file.name
    ext = os.path.basename(fileN).split('.')[1]
df = None
if file is not None:
    if ext !='xlsx':
        st.write('Unsupported file format, first save the excel as xlsx and try again')
        st.stop()
    else:
        df = pd.read_excel(file)
        st.write('Excel accepted')
    if df is not None:
        columns = ['A', 'AS', 'VD', 'RD','TO', 'TI', 'DD']
        cols = df.columns.to_list()
        if not all(column in cols for column in columns):
            missing_columns = [column for column in columns if column not in cols]
            for column in missing_columns:
                st.markdown(f' **ERROR !!! {column} is not in the file uploaded**')
                st.markdown('**First rename all the columns as guided above**')
                st.stop()
        else:
              # Convert 'A' column to string and create 'ART' column with numeric part
            df['A'] = df['A'].astype(str)
            df['ART'] = df['A'].str.replace('[^0-9]', '', regex=True)
            df['ART'] = pd.to_numeric(df['ART'], errors= 'coerce')
            df = df[df['ART']>0]
            #df.dropna(subset='ART', inplace=True)
            df = df.copy()
            df[['AS', 'RD', 'VD','TO','TI']] = df[['AS', 'RD', 'VD','TO','TI']].astype(str)
            if df['TI'].str.contains('YES').any():
                st.write("The transfer in column you are using doesn't have dates but words, like YES, kindly use the right transfer in colum")
                st.stop()
            
            df['AS'] = df['AS'].astype(str)
            df['RD'] = df['RD'].astype(str)
            df['TI'] = df['TI'].astype(str)
            df['TO'] = df['TO'].astype(str)
            df['VD'] = df['VD'].astype(str)
            df['DD'] = df['DD'].astype(str)
            
            y = pd.DataFrame({'A' :['2','3','4'], 'TI':['1-1-1',1,'1/1/1'], 'RD':['1-1-1',1,'1/1/1'],'DD':['1-1-1',1,'1/1/1'], 
                              'TO':['1-1-1',1,'1/1/1'], 'AS':['1-1-1',1,'1/1/1'], 'VD':['1-1-1',1,'1/1/1']})  
            

            df['AS'] = df['AS'].str.replace('00:00:00', '', regex=True)
            df['RD'] = df['RD'].astype(str)
            df['RD'] = df['RD'].str.replace('00:00:00', '', regex=True)
            df['DD'] = df['DD'].str.replace('00:00:00', '', regex=True)
            df['VD'] = df['VD'].str.replace('00:00:00', '', regex=True)
            df['TO'] = df['TO'].str.replace('00:00:00', '', regex=True)
            df['TI'] = df['TI'].str.replace('00:00:00', '',regex=True)
            df = pd.concat([df,y])


            df['AS'] = df['AS'].astype(str)
            df['RD'] = df['RD'].astype(str)
            df['TI'] = df['TI'].astype(str)
            df['TO'] = df['TO'].astype(str)
            df['VD'] = df['VD'].astype(str)
            df['DD'] = df['DD'].astype(str)


            # SPLITTING ART START DATE
            A = df[df['AS'].str.contains('-')]
            a = df[~df['AS'].str.contains('-')]
            B = a[a['AS'].str.contains('/')]
            C = a[~a['AS'].str.contains('/')]

            A[['Ayear', 'Amonth', 'Aday']] = A['AS'].str.split('-', expand = True)
            B[['Ayear', 'Amonth', 'Aday']] = B['AS'].str.split('/', expand = True)
                        
            C['AS'] = pd.to_numeric(C['AS'], errors='coerce')
            C['AS'] = pd.to_datetime(C['AS'], origin='1899-12-30', unit='D', errors='coerce')
            C['AS'] =  C['AS'].astype(str)
            C[['Ayear', 'Amonth', 'Aday']] = C['AS'].str.split('-', expand = True)
            df = pd.concat([A,B,C])

             # SPLITTING DEATH DATE
            A = df[df['DD'].str.contains('-')]
            a = df[~df['DD'].str.contains('-')]
            B = a[a['DD'].str.contains('/')]
            C = a[~a['DD'].str.contains('/')]

            A[['Dyear', 'Dmonth', 'Dday']] = A['DD'].str.split('-', expand = True)
            B[['Dyear', 'Dmonth', 'Dday']] = B['DD'].str.split('/', expand = True)
                        
            C['DD'] = pd.to_numeric(C['DD'], errors='coerce')
            C['DD'] = pd.to_datetime(C['DD'], origin='1899-12-30', unit='D', errors='coerce')
            C['DD'] =  C['DD'].astype(str)
            C[['Dyear', 'Dmonth', 'Dday']] = C['DD'].str.split('-', expand = True)
            df = pd.concat([A,B,C])
          
            # SORTING THE RETURN VISIT DATE
            A = df[df['RD'].str.contains('-')].copy()
            a = df[~df['RD'].str.contains('-')].copy()
            B = a[a['RD'].str.contains('/')].copy()
            C = a[~a['RD'].str.contains('/')].copy()
      
            A[['Ryear', 'Rmonth', 'Rday']] = A['RD'].str.split('-', expand = True)
            B[['Ryear', 'Rmonth', 'Rday']] = B['RD'].str.split('/', expand = True)
                        
            C['RD'] = pd.to_numeric(C['RD'], errors='coerce')
            C['RD'] = pd.to_datetime(C['RD'], origin='1899-12-30', unit='D', errors='coerce')
            C['RD'] =  C['RD'].astype(str)
            C[['Ryear', 'Rmonth', 'Rday']] = C['RD'].str.split('-', expand = True)
            df = pd.concat([A,B,C]) 
          
            #SORTING THE VD DATE
            A = df[df['VD'].str.contains('-')].copy()
            a = df[~df['VD'].str.contains('-')].copy()
            B = a[a['VD'].str.contains('/')].copy()
            C = a[~a['VD'].str.contains('/')].copy()

            A[['Vyear', 'Vmonth', 'Vday']] = A['VD'].str.split('-', expand = True)
            B[['Vyear', 'Vmonth', 'Vday']] = B['VD'].str.split('/', expand = True)
                        
            C['VD'] = pd.to_numeric(C['VD'], errors='coerce')
            C['VD'] = pd.to_datetime(C['VD'], origin='1899-12-30', unit='D', errors='coerce')
            C['VD'] =  C['VD'].astype(str)
            C[['Vyear', 'Vmonth', 'Vday']] = C['VD'].str.split('-', expand = True)
            df = pd.concat([A,B,C])

            #SORTING THE TO DATE
            A = df[df['TO'].str.contains('-')].copy()
            a = df[~df['TO'].str.contains('-')].copy()
            B = a[a['TO'].str.contains('/')].copy()
            C = a[~a['TO'].str.contains('/')].copy()

            A[['Tyear', 'Tmonth', 'Tday']] = A['TO'].str.split('-', expand = True)
            B[['Tyear', 'Tmonth', 'Tday']] = B['TO'].str.split('/', expand = True)
                        
            C['TO'] = pd.to_numeric(C['TO'], errors='coerce')
            C['TO'] = pd.to_datetime(C['TO'], origin='1899-12-30', unit='D', errors='coerce')
            C['TO'] =  C['TO'].astype(str)
            C[['Tyear', 'Tmonth', 'Tday']] = C['TO'].str.split('-', expand = True)
            df = pd.concat([A,B,C])
        

           #SORTING THE TI DATE
            A = df[df['TI'].str.contains('-')].copy()
            a = df[~df['TI'].str.contains('-')].copy()
            B = a[a['TI'].str.contains('/')].copy()
            C = a[~a['TI'].str.contains('/')].copy()

            A[['Tiyear', 'Timonth', 'Tiday']] = A['TI'].str.split('-', expand = True)
            B[['Tiyear', 'Timonth', 'Tiday']] = B['TI'].str.split('/', expand = True)
                        
            C['TI'] = pd.to_numeric(C['TI'], errors='coerce')
            C['TI'] = pd.to_datetime(C['TI'], origin='1899-12-30', unit='D', errors='coerce')
            C['TI'] =  C['TI'].astype(str)
            C[['Tiyear', 'Timonth', 'Tiday']] = C['TI'].str.split('-', expand = True)
            df = pd.concat([A,B,C])

               #BRINGING BACK THE / IN DATES
            #df[['AS', 'RD', 'VD','TO','TI']] = df[['AS', 'RD', 'VD','TO','TI']].astype(str)
            df['AS'] = df['AS'].astype(str)
            df['RD'] = df['RD'].astype(str)
            df['TI'] = df['TI'].astype(str)
            df['TO'] = df['TO'].astype(str)
            df['VD'] = df['VD'].astype(str)
            df['DD'] = df['DD'].astype(str)

            #Clearing NaT from te dates
            df['AS'] = df['AS'].str.replace('NaT', '',regex=True)
            df['RD'] = df['RD'].str.replace('NaT', '',regex=True)
            df['VD'] = df['VD'].str.replace('NaT', '',regex=True)
            df['TO'] = df['TO'].str.replace('NaT', '',regex=True)
            df['TI'] = df['TI'].str.replace('NaT', '',regex=True)
            df['DD'] = df['DD'].str.replace('NaT', '',regex=True)

            #             #SORTING THE VIRAL LOAD YEARS
          
            df[['Vyear', 'Vmonth', 'Vday']] =df[['Vyear', 'Vmonth', 'Vday']].apply(pd.to_numeric, errors = 'coerce') 
            df['Vyear'] = df['Vyear'].fillna(994)
            a = df[df['Vyear']>31].copy()
            b = df[df['Vyear']<32].copy()
            b = b.rename(columns={'Vyear': 'Vday2', 'Vday': 'Vyear'})
            b = b.rename(columns={'Vday2': 'Vday'})
            df = pd.concat([a,b])
            dfa = df.shape[0]


            try:
               df[['Tiyear', 'Tiday']] =df[['Tiyear','Tiday']].apply(pd.to_numeric, errors = 'coerce')
            except:
                st.write('**There are no dates in the transfer in column**')
                #st. markdown('##')
                st.write('Copy one date from the Return Visit date and paste it in the Transfer in date, and try again')
                st.write('But this will mean the number of Transfer in is wrong but other paarameters will be correct')
                st.markdown('##')
                st.write('**Another option is to extract a new extract with Transfer in Obs date**')
                st.stop()
            df['Tiyear'] = df['Tiyear'].fillna(994)
            a = df[df['Tiyear']>31].copy()
            b = df[df['Tiyear']<32].copy()
            b = b.rename(columns={'Tiyear': 'Tiday2', 'Tiday': 'Tiyear'})
            b = b.rename(columns={'Tiday2': 'Tiday'})
            df = pd.concat([a,b])
            dfb = df.shape[0]

            # #SORTING THE RETURN VISIT DATE YEARS
            #st.write(df['RD'])
            #st.stop()
            df[['Rday', 'Ryear']] = df[['Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
            
            df['Ryear'] = df['Ryear'].fillna(994)
            a = df[df['Ryear']>31].copy()
            b = df[df['Ryear']<32].copy()
            b = b.rename(columns={'Ryear': 'Rday2', 'Rday': 'Ryear'})
            b = b.rename(columns={'Rday2': 'Rday'})

            df = pd.concat([a,b])
            dfc = df.shape[0]
            
                #SORTING THE TRANSFER OUT DATE YEAR
            df[['Tday', 'Tyear']] = df[['Tday', 'Tyear']].apply(pd.to_numeric, errors='coerce')
            df['Tyear'] = df['Tyear'].fillna(994)
            a = df[df['Tyear']>31].copy()
            b = df[df['Tyear']<32].copy()
            b = b.rename(columns={'Tyear': 'Tday2', 'Tday': 'Tyear'})
            b = b.rename(columns={'Tday2': 'Tday'})
            df = pd.concat([a,b])

            
               #SORTING THE ART START YEARS
            df[['Ayear', 'Amonth', 'Aday']] =df[['Ayear', 'Amonth', 'Aday']].apply(pd.to_numeric, errors = 'coerce')
            df['Ayear'] = df['Ayear'].fillna(994)
            a = df[df['Ayear']>31].copy()
            b = df[df['Ayear']<32].copy()
            b = b.rename(columns={'Ayear': 'Aday2', 'Aday': 'Ayear'})
            b = b.rename(columns={'Aday2': 'Aday'})
            df = pd.concat([a,b])
            dfe = df.shape[0]

              #SORTING THE ART START YEARS
            df[['Dyear', 'Dmonth', 'Dday']] =df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors = 'coerce')
            df['Dyear'] = df['Dyear'].fillna(994)
            a = df[df['Dyear']>31].copy()
            b = df[df['Dyear']<32].copy()
            b = b.rename(columns={'Dyear': 'Dday2', 'Dday': 'Dyear'})
            b = b.rename(columns={'Dday2': 'Dday'})
            df = pd.concat([a,b])
            dfe = df.shape[0]
           
            #file = r"C:\Users\Desire Lumisa\Desktop\TX CURR\MATEETE.xlsx"
            file2 = r'ALL.xlsx'
            dfx = pd.read_excel(file2)
            
            df[['Tyear', 'Ryear', 'Rmonth', 'Rday', 'Vyear', 'Vmonth', 'Ayear']] = df[['Tyear', 'Ryear', 'Rmonth', 'Rday', 'Vyear', 'Vmonth', 'Ayear']].apply(pd.to_numeric, errors='coerce')
            oneyear = df.copy()
          
            dfw = df[df['Ryear'] ==2025].copy()
            dfy = df[df['Ryear'] ==2024].copy()
            dfy[['Rmonth', 'Rday']] = dfy[['Rmonth', 'Rday']].apply(pd.to_numeric, errors = 'coerce')
            dfy = dfy[((dfy['Rmonth']>6) | ((dfy['Rmonth']==6) & (dfy['Rday'] >2)))].copy()
            df = pd.concat([dfw,dfy])
            potential = df.shape[0]
            dpot = df.copy()
            
            df['Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
            dead = df[df['Dyear']!=994].copy()

            df = df[df['Dyear']==994].copy()
            
            df['Rday1'] = df['Rday'].astype(str).str.split('.').str[0]
            df['Rmonth1'] = df['Rmonth'].astype(str).str.split('.').str[0]
            df['Ryear1'] = df['Ryear'].astype(str).str.split('.').str[0]

            df['Vday1'] = df['Vday'].astype(str).str.split('.').str[0]
            df['Vmonth1'] = df['Vmonth'].astype(str).str.split('.').str[0]
            df['Vyear1'] = df['Vyear'].astype(str).str.split('.').str[0]

            #df['Tiday'] = df['Tiday'].astype(str).str.split('.').str[0]
            #df['Timonth'] = df['Timonth'].astype(str).str.split('.').str[0]
            #df['Tiyear'] = df['Tiyear'].astype(str).str.split('.').str[0]

            df['Aday1'] = df['Aday'].astype(str).str.split('.').str[0]
            df['Amonth1'] = df['Amonth'].astype(str).str.split('.').str[0]
            df['Ayear1'] = df['Ayear'].astype(str).str.split('.').str[0]
            
            df['Tday1'] = df['Tday'].astype(str).str.split('.').str[0]
            df['Tmonth1'] = df['Tmonth'].astype(str).str.split('.').str[0]
            df['Tyear1'] = df['Tyear'].astype(str).str.split('.').str[0]

            df['ART START DATE'] = df['Aday1'] + '/' + df['Amonth1'] + '/' + df['Ayear1']
            df['RETURN DATE'] = df['Rday1'] + '/' + df['Rmonth1'] + '/' + df['Ryear1']
            df['VL DATE'] = df['Vday1'] + '/' + df['Vmonth1'] + '/' + df['Vyear1']
            df['T OUT DATE'] = df['Tday1'] + '/' + df['Tmonth1'] + '/' + df['Tyear1']
            #df['T IN DATE'] = df['Rday1'] + '/' + df['Rmonth1'] + '/' + df['Ryear1']

            df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'], format='%d/%m/%Y', errors='coerce')
            df['VL DATE'] = pd.to_datetime(df['VL DATE'], format='%d/%m/%Y', errors='coerce')
            df['T OUT DATE'] = pd.to_datetime(df['T OUT DATE'], format='%d/%m/%Y', errors='coerce')
            df['ART START DATE'] = pd.to_datetime(df['ART START DATE'], format='%d/%m/%Y', errors='coerce')

            df['RETURN DATE'] = df['RETURN DATE'].dt.strftime('%d/%m/%Y')
            df['VL DATE'] = df['VL DATE'].dt.strftime('%d/%m/%Y')
            df['T OUT DATE'] = df['T OUT DATE'].dt.strftime('%d/%m/%Y')
            df['ART START DATE'] = df['ART START DATE'].dt.strftime('%d/%m/%Y')
            
            df = df.rename(columns={'A': 'ART NO'})#, 'AS': 'ART START DATE', 'RD': 'RETURN DATE', 'VD': 'VL DATE', 'TO': 'T OUT DATE'})
            
            df[['Tyear', 'Ryear', 'Rmonth', 'Rday', 'Vyear', 'Vmonth', 'Ayear']] = df[['Tyear', 'Ryear', 'Rmonth', 'Rday', 'Vyear', 'Vmonth', 'Ayear']].apply(pd.to_numeric, errors='coerce')
            
            
            TXML = df[df['Ryear']==2024].copy()
            TXML[['Rmonth', 'Rday']] = TXML[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
            TXML = TXML[((TXML['Rmonth']>6) | ((TXML['Rmonth']==6) & (TXML['Rday']>2)))].copy()
            TXML[['Rmonth', 'Rday']] = TXML[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
            TXML = TXML[((TXML['Rmonth']<9) | ((TXML['Rmonth']==9) & (TXML['Rday']<3)))].copy()
            TXML['Tyear'] = pd.to_numeric(TXML['Tyear'], errors='coerce')
            TXML = TXML[TXML['Tyear']==994].copy()
            
            #TX CURR
            df['Ryear'] = pd.to_numeric(df['Ryear'], errors='coerce')
            a = df[df['Ryear']==2025].copy()
            a['Tyear'] = pd.to_numeric(a['Tyear'], errors='coerce')
            a = a[a['Tyear']==994].copy()
            
            b = df[df['Ryear']==2024].copy()
            b[['Rmonth', 'Rday']] = b[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
            b = b[((b['Rmonth']>9) | ((b['Rmonth']==9) & (b['Rday']>2)))].copy()
            b['Tyear'] = pd.to_numeric(b['Tyear'], errors='coerce')
            b = b[b['Tyear']==994].copy()
            TXCURR = pd.concat([a,b])
            
            #TX NEW
            df[['Ayear', 'Amonth']] = df[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
            TXNEW = df[((df['Ayear']==2024) & (df['Amonth'].isin([7,8,9])))].copy()
            df[['Tiyear', 'Timonth']] = df[['Tiyear', 'Timonth']].apply(pd.to_numeric, errors='coerce')
            TI = df[((df['Tiyear']==2024) & (df['Timonth'].isin([7,8,9])))].copy()
            

            #TRANSFER OUTS
            df[['Tyear', 'Tmonth']] = df[['Tyear', 'Tmonth']].apply(pd.to_numeric, errors='coerce')
            TO = df[df['Tyear']!=994].copy()
            TO[['Ryear', 'Rmonth', 'Rday']] = TO[['Ryear', 'Rmonth','Rday']].apply(pd.to_numeric, errors='coerce')
            
            TOa = TO[((TO['Ryear']==2024) & (TO['Rmonth']<10))].copy()
            TOa[['Rmonth', 'Rday']] = TOa[['Rmonth','Rday']].apply(pd.to_numeric, errors='coerce')
            #st.write(TOa)
            TOa = TOa[((TOa['Rmonth'] >6) | ((TOa['Rmonth'] ==6) & (TOa['Rday'] >2)))].copy()
            #TOa[['Tmonth', 'Tyear']] = TOa[['Tmonth','Tyear']].apply(pd.to_numeric, errors='coerce')
            #TOa = TOa[((TOa['Tyear']==2024) & (TOa['Tmonth'].isin([4,5,6])))].copy()

            TO[['Ryear', 'Rmonth', 'Rday']] = TO[['Ryear', 'Rmonth','Rday']].apply(pd.to_numeric, errors='coerce')
            FALSE = TO[((TO['Ryear']>2024) | ((TO['Ryear']==2024) & (TO['Rmonth']>9)))].copy()

            TXCUR = pd.concat([TXCURR,FALSE])

            #VL COV
            TXCUR['Ayear'] = pd.to_numeric(TXCUR['Ayear'], errors='coerce')
            c = TXCUR[ TXCUR['Ayear']==2024].copy()
            d = TXCUR[ TXCUR['Ayear']<2024].copy()
            d[['Vyear', 'Vmonth']] = d[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
            e = d[((d['Vyear'] ==2024) | ((d['Vyear'] ==2023) & (d['Vmonth'] >9)))].copy()
            f = d[((d['Vyear'] < 2023) | ((d['Vyear'] ==2023) & (d['Vmonth'] <10)))].copy()
            WVL = pd.concat([c,e])
            
            NOVL = f.copy()
            cphl = r'AVLS.csv'
            cp = pd.read_csv(cphl)

            POTENTIAL = potential
            newad = TXNEW.shape[0]
            out = TOa.shape[0]
            inn = TI.shape[0]
            curr = TXCUR.shape[0]
            false = FALSE.shape[0]
            lost = TXML.shape[0]
            vl = WVL.shape[0]
            #st.write(vl)
            #st.write(curr)
            #st.stop()
            perc = round((vl/curr)*100)
            exp = round(curr*0.95)
            novl = NOVL.shape[0]
            current_time = time.localtime()
            week = time.strftime("%V", current_time)
            week = int(week) +13

            districts = list(dfx['DISTRICT'].unique())
            district = st.radio(label='**Choose a district**', options=districts,index=None, horizontal=True)
            if district:
                facilities = dfx[dfx['DISTRICT']==district]
                facilities = facilities['FACILITY']
                facility = st.selectbox(label='**Choose a facility**', options=facilities,index=None)
                if facility:
                    preva = dfx[dfx['FACILITY'] == facility]
                    prev = int(preva.iloc[0,3])
                    name =str(preva.iloc[0,4])
                    UK = potential- prev - inn - newad
                    dd = dead.shape[0]

                    ba = prev - curr
                    if ba > 0:
                        bal = ba
                    elif ba == 0:
                        bal = 'EVEN'
                    elif ba < 0:
                        bal = 'EXCEEDED'
                    grow = curr-prev
                    if grow ==0:
                        st.success(f'WEBALE {name},😐 this TXCURR has broken even (Q3 CURR is equal to Q4 CURR), but you need to add more clients to grow it even further 👏👏👏')
                        if perc > 94:
                            st.success(f'Even the VL COVERAGE is good, at {perc}%  👏👏👏')
                        else:
                            st.warning(f'**However the VL COVERAGE is poor, at {perc}%** 🥲')

                    elif grow>0:
                        st.success(f'WEBALE {name},😐 you have grown this TXCURR by {grow}, but you need to audit the TIs and TXNEWs, and watch out for RTT 👏👏👏')
                        if perc > 94:
                            st.success(f'Even the VL COVERAGE is good, at {perc}%  👏👏👏')
                            st.balloons()
                            time.sleep(2)
                            st.balloons()
                            time.sleep(2)
                            st.balloons()
                            time.sleep(2)
                            st.balloons()
                            time.sleep(2)
                            
                        else:
                            st.warning(f'**However the VL COVERAGE is poor, at {perc}%** 🥲')
                            st.balloons()
                            time.sleep(2)
                            st.balloons()                       
                    else:
                        st.warning(f'**BANANGE {name}, 😢 you have dropped this TXCURR by {grow}, you need to audit the TXMLs and TOs, and watch out for the dead** 😢😢😢')
                        if perc > 94:
                            st.success(f'BUT the VL COVERAGE is good, at {perc}% 👏')
                        else:
                            st.warning(f'**EVEN the VL COVERAGE is poor, at {perc}%** 😢😢😢')
                    cp = cp[cp['facility']==facility].copy()
                    cp = cp.rename(columns ={'ART-NUMERIC': 'ART'})
                    cp['ART'] = pd.to_numeric(cp['ART'], errors='coerce')
                    NOV = NOVL[['ART', 'RETURN DATE', 'VL DATE']].copy()
                    NOV['ART'] = pd.to_numeric(NOV['ART'], errors='coerce')

                    AT = pd.merge(cp, NOV, on='ART',how='inner')
                    NOVL['ART'] = pd.to_numeric(NOVL['ART'], errors='coerce')
                    
                    AT['ART'] = pd.to_numeric(AT['ART'], errors='coerce')
                    TRUE = NOVL[~NOVL['ART'].isin(AT['ART'])].copy()
                    
                    oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                    new = oneyear[((oneyear['Ayear']==2023) & (oneyear['Amonth'].isin([7,8,9])))].copy()
    
                    new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                    tin = new[new['Tiyear']!=994].copy()
                    #one =new.shape[0]
                    tew = tin.shape[0]
                    
                    newtotal = new.shape[0]
                    
                    new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
                    newdead = new[new['Dyear']!=994].copy()

                    deadnew = newdead.shape[0]
                    new = new[new['Dyear']==994].copy()

                    new['Tyear'] = pd.to_numeric(new['Tyear'], errors='coerce')
                    
                    newto = new[new['Tyear']!=994].copy()
                    outnew = newto.shape[0]
                    
                    new = new[new['Tyear']==994].copy()
                    new['ART'] = pd.to_numeric(new['ART'], errors = 'coerce')
                    TXCUR['ART'] = pd.to_numeric(TXCUR['ART'], errors = 'coerce')
                    
                    active = new[new['ART'].isin(TXCUR['ART'])].copy()
                    lostn = new[~new['ART'].isin(TXCUR['ART'])].copy()
                    

                    newactive = active.shape[0]
                    newlost = lostn.shape[0]
                    #st.write(newlost)
                           
                    #ret = newtotal - newlost
                    if newtotal == 0:
                        rete = 0
                    elif newactive == 0:
                        rete = 0
                    else:
                        rete = round((newactive/newtotal)*100)
                        rete = f"{rete} %"
                    
                    data = pd.DataFrame([{
                                'DISTRICT': district,
                                'FACILITY' : facility,
                                'Q3 CURR':prev,
                                'UNKNOWN GAIN': UK,
                                'DEAD': dd,
                                'POTENTIAL': potential,
                                'Q4 CURR': curr,
                                'TXML' : lost,
                                 'BALANCE': bal,
                                'TX NEW' : newad,
                                'TO' : out,
                                'FALSE TO': false,
                                'TI': inn,
                                'HAS VL' : vl,
                                'VL COV (%)': perc,
                                'EXPECTED': exp,
                                'NO VL' : novl,
                                'WEEK': week,
                                 'ORIGINAL COHORT': newtotal,
                                 'ONE YEAR TI': tew,
                                 'ONE YEAR LOST': newlost,
                                  'ONE YEAR TO': outnew,
                                 'ONE YEAR DEAD': deadnew,
                                 'ONE YEAR ACTIVE': newactive,
                                 'ONE YR RETENTION': rete
                                 }])
                    #data = data.set_index('DISTRICT')
                    
                    #SUBMISSION
                    # conn = st.connection('gsheets', type=GSheetsConnection)
                    # exist = conn.read(worksheet ='TXML', usecols = list(range(15)), ttl=5)
                    # existing = exist.dropna(how='all')
                    col1,col2,col3 = st.columns([1,2,1])
                    with col3:
                        submit = st.button('Submit') 
                    if submit:
                        try:
                            conn = st.connection('gsheets', type=GSheetsConnection)
                            exist = conn.read(worksheet ='MPIGI', usecols = list(range(25)), ttl=5)
                            existing = exist.dropna(how='all')
                            updated = pd.concat([existing, data], ignore_index =True)
                            conn.update(worksheet = 'TXML', data = updated)
                            st.success('Your data above has been submitted')
                        except:
                            st.write("Couldn't submit, poor network")
                    if submit:
                        st.write('**TX CURR AS OF 3rd SEPT**')
                        pass
                        st.session_state.submited = True
                    else:
                        st.write('')
                    if st.session_state.submited:
                        st.dataframe(data)
                        st.write(f"<h6>DOWNLOAD LINELISTS FROM HERE</h6>", unsafe_allow_html=True)
                        cola, colb, colc = st.columns(3)
                        with cola:
                             dat = TXML.copy()
                             
                             dat = dat[['ART NO', 'ART START DATE', 'RETURN DATE', 'VL DATE']]
                             csv_data = dat.to_csv(index=False)
                             st.download_button(
                                         label=" DOWNLOAD TXML",
                                         data=csv_data,
                                         file_name=f"{facility} TXML.csv",
                                         mime="text/csv")
                        with colb:
                             dat = NOVL.copy()
                             dat = dat[['ART NO', 'ART START DATE', 'RETURN DATE', 'VL DATE']]
                             
                             csv_data = dat.to_csv(index=False)
                             st.download_button(
                                             label=" DOWNLOAD WITH NO VL",
                                             data=csv_data,
                                             file_name=f" {facility} NO VL.csv",
                                             mime="text/csv")
                        with colc:
                             dat = TOa.copy()
                             dat = dat[['ART NO', 'ART START DATE', 'RETURN DATE', 'VL DATE', 'T OUT DATE']]
                             dat = AT.copy()
                             csv_data = dat.to_csv(index=False)
                             st.download_button(
                                         label=" DOWNLOAD TRANSFER OUTS",
                                         data=csv_data,
                                         file_name=f" {facility} T OUTS.csv",
                                         mime="text/csv")
    
    #########################################################################################################################################################
                        AT= AT[['ART', 'RETURN DATE', 'VL DATE','art_number','date_collected','result_numeric']].copy()
                        a = AT.shape[0]
                        if a==0:   
                            st.write('**I DO NOT SEE VL RESULTS AT CPHL MISSING IN THIS EMR EXTRACT FOR NOW. WAIT FOR FUTURE UPDATES**')
                        elif a == 1:
                            st.success(f'I see only **{a}** result at CPHL that is not yet entered into EMR')
                        else:
                            st.success(f'I see over **{a}** results at CPHL that are not yet entered into EMR')
                                
            
                        cola, colb = st.columns([2,1])
                        with cola:
                            if a>0:  
                                named = facility
                                #if st.button('DOWNLOAD FILE'):
                                wb = Workbook()
                                ws = wb.active
                     
                        # Convert DataFrame to Excel
                                for r_idx, row in enumerate(AT.iterrows(), start=1):
                                       for c_idx, value in enumerate(row[1], start=1):
                                            ws.cell(row=r_idx, column=c_idx, value=value)
                                ws.insert_rows(0,2)
                    
                                blue = PatternFill(fill_type = 'solid', start_color = 'F6F8F7')
                                    # ws.column_dimensions['H'].width = 14
                    
                                for num in range (1, ws.max_row+1):
                                     for letter in ['D','E', 'F']:
                                          ws[f'{letter}{num}'].font = Font(b= True, i = True)
                                          ws[f'{letter}{num}'].font = Font(b= True, i = True)
                                          ws[f'{letter}{num}'].fill = blue
                                          ws[f'{letter}{num}'].border = Border(top = Side(style = 'thin', color ='000000'),
                                                                                right = Side(style = 'thin', color ='000000'),
                                                                                left = Side(style = 'thin', color ='000000'),
                                                                                bottom = Side(style = 'thin', color ='000000'))
                                ws['B1'] ='EMR DETAILS'
                                ws['F1'] = 'CPHL DETAILS'
                                ws['A2'] = 'ART-NO'
                                ws['B2'] = 'RETUR VISIT DATE'
                                ws['C2'] = 'EMR VL DATE' 
                                ws['D2'] = 'ART NO'
                                ws['E2'] = 'CPHL DATE'
                                ws['F2']  = 'CPHL RESULTS'
                    
                    
                                letters = ['B', 'C', 'D','E','F']
                                for letter in letters:
                                      ws.column_dimensions[letter].width =15
                    
                                ran = random.random()
                                rand = round(ran,2)
                                file_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'{named}_missing_results {rand}.xlsx')
                                directory = os.path.dirname(file_path)
                                Path(directory).mkdir(parents=True, exist_ok=True)
                    
                                      # Save the workbook
                                wb.save(file_path)
                         # Serve the file for download
                                with open(file_path, 'rb') as f:
                                  file_contents = f.read()           
                                st.download_button(label=f'DONLOAD MISSING RESULTS FOR {named} ', data=file_contents,file_name=f'{named}_missing_results {rand}.xlsx', 
                                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                            else:
                                pass
                        with colb:
                             if dd >0:             
                                 st.write('**CONFIRM IF THESE ARE DEAD**')
                                 dat = dead.copy()
                                 #st.write(dead.columns)
                                 dat = dat[['A', 'RD', 'DD']]
                                 dat = AT.copy()
                                 csv_data = dat.to_csv(index=False)
                                 st.download_button(
                                             label=" DOWNLOAD DEAD LINELIST",
                                             data=csv_data,
                                             file_name=f" {facility} DEAD.csv",
                                             mime="text/csv")
                             else:
                                 pass
                                
    #########################################################################################################
                        #st.write(lostn)
                        st.write('**ONE YEAR COHORT RETENTION**')
                        one = data[['ORIGINAL COHORT','ONE YEAR LOST','ONE YEAR DEAD', 'ONE YEAR TO' ,'ONE YEAR ACTIVE',  'ONE YR RETENTION']].copy()
                                 
                        one = one.rename(columns ={'ONE YEAR LOST': 'LOST','ONE YEAR DEAD': 'DEAD','ONE YEAR TO':'TO',  'ONE YEAR ACTIVE': 'ACTIVE'})
                        one = one.set_index('ORIGINAL COHORT')
                        st.write(one)
                        cola,colb = st.columns(2)
                        if newlost==0:
                            st.write('**NO IIT AMONGST LOST CLIENTS, SO NO LINE LIST TO DOWNLOAD**')
                            pass
                        else:         
                            lostn = lostn.rename(columns ={'A':'ART NO', 'AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VL DATE'})
                            with cola:
                                 dat = lostn.copy()
                                 #dat = TXCUR.copy()
                                 dat = dat[['ART NO', 'ART START DATE','RETURN DATE', 'TI']]
                                 csv_data = dat.to_csv(index=False)
                                 st.download_button(
                                             label="DOWNLOAD_IIT_FOR_1_YR_COHORT",
                                             data=csv_data,
                                             file_name=f"{facility}_1_YR_IIT.csv",
                                             mime="text/csv")
    
                        if   outnew ==0:
                            pass
                        else:
                             with colb:
                                 dat =  newto.copy()
                                 #st.write(dat.columns)
                                 #dat = TXCUR.copy()
                                 dat = dat[['A', 'AS','RD', 'TO','TI']]
                                 csv_data = dat.to_csv(index=False)
                                 st.download_button(
                                             label="DOWNLOAD_TOs_FOR_1_YR_COHORT",
                                             data=csv_data,
                                             file_name=f"{facility}_1_YR_TOs.csv",
                                             mime="text/csv")
                    else:
                        st.write('**FIRST SUBMIT TO SEE THE LINE-LISTS**')
                        st.stop()
                        

                                 
                                
                    
